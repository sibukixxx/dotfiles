#!/usr/bin/env python3
"""
SQLite to Excel Converter - Export SQLite database tables to Excel spreadsheet.

Usage:
    python sqlite_to_xlsx.py <database.db> [options]

Options:
    --output FILE          Output Excel file (default: <database>_export.xlsx)
    --tables TABLE,...     Comma-separated list of tables to export (default: all)
    --query "SQL"          Export result of custom SQL query
    --query-name NAME      Sheet name for custom query result (default: Query_Result)
    --include-schema       Add a sheet with table schemas
    --format               Apply basic formatting (headers, auto-width)
"""

import argparse
import sqlite3
import sys
from pathlib import Path

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.utils import get_column_letter


def get_tables(conn: sqlite3.Connection) -> list[str]:
    """Get all table names from database."""
    cursor = conn.execute(
        "SELECT name FROM sqlite_master WHERE type='table' AND name NOT LIKE 'sqlite_%'"
    )
    return [row[0] for row in cursor.fetchall()]


def get_table_schema(conn: sqlite3.Connection, table: str) -> list[dict]:
    """Get schema information for a table."""
    cursor = conn.execute(f"PRAGMA table_info({table})")
    columns = []
    for row in cursor.fetchall():
        columns.append({
            "cid": row[0],
            "name": row[1],
            "type": row[2],
            "notnull": bool(row[3]),
            "default": row[4],
            "pk": bool(row[5]),
        })
    return columns


def get_row_count(conn: sqlite3.Connection, table: str) -> int:
    """Get row count for a table."""
    cursor = conn.execute(f"SELECT COUNT(*) FROM {table}")
    return cursor.fetchone()[0]


def export_table_to_dataframe(conn: sqlite3.Connection, table: str) -> pd.DataFrame:
    """Export a table to pandas DataFrame."""
    return pd.read_sql_query(f"SELECT * FROM {table}", conn)


def execute_query(conn: sqlite3.Connection, query: str) -> pd.DataFrame:
    """Execute custom SQL query and return DataFrame."""
    return pd.read_sql_query(query, conn)


def apply_formatting(ws, df: pd.DataFrame):
    """Apply professional formatting to worksheet."""
    # Header style
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="4472C4")
    header_alignment = Alignment(horizontal="center", vertical="center")

    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    # Format headers (row 1)
    for col_idx, col_name in enumerate(df.columns, 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border

    # Format data cells and apply borders
    for row_idx in range(2, len(df) + 2):
        for col_idx in range(1, len(df.columns) + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.border = thin_border

    # Auto-adjust column widths
    for col_idx, col_name in enumerate(df.columns, 1):
        max_length = len(str(col_name))
        for row_idx in range(2, min(len(df) + 2, 102)):  # Sample first 100 rows
            cell_value = ws.cell(row=row_idx, column=col_idx).value
            if cell_value:
                max_length = max(max_length, len(str(cell_value)))

        adjusted_width = min(max_length + 2, 50)  # Cap at 50
        ws.column_dimensions[get_column_letter(col_idx)].width = adjusted_width

    # Freeze header row
    ws.freeze_panes = "A2"


def create_schema_sheet(wb: Workbook, conn: sqlite3.Connection, tables: list[str]):
    """Create a sheet with schema information for all tables."""
    ws = wb.create_sheet("_Schema")

    # Headers
    headers = ["Table", "Column", "Type", "NotNull", "Default", "PrimaryKey"]
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="70AD47")

    row_idx = 2
    for table in tables:
        schema = get_table_schema(conn, table)
        for col_info in schema:
            ws.cell(row=row_idx, column=1, value=table)
            ws.cell(row=row_idx, column=2, value=col_info["name"])
            ws.cell(row=row_idx, column=3, value=col_info["type"])
            ws.cell(row=row_idx, column=4, value="YES" if col_info["notnull"] else "NO")
            ws.cell(row=row_idx, column=5, value=col_info["default"])
            ws.cell(row=row_idx, column=6, value="YES" if col_info["pk"] else "NO")
            row_idx += 1

    # Auto-width
    for col_idx in range(1, 7):
        ws.column_dimensions[get_column_letter(col_idx)].width = 15

    ws.freeze_panes = "A2"


def sqlite_to_xlsx(
    db_path: Path,
    output_path: Path,
    tables: list[str] = None,
    custom_query: str = None,
    query_name: str = "Query_Result",
    include_schema: bool = False,
    apply_format: bool = True,
) -> dict:
    """Convert SQLite database to Excel file."""

    conn = sqlite3.connect(db_path)
    result = {
        "success": True,
        "output_file": str(output_path),
        "sheets": [],
        "total_rows": 0,
    }

    try:
        wb = Workbook()
        wb.remove(wb.active)  # Remove default sheet

        # Get tables to export
        all_tables = get_tables(conn)
        if tables:
            export_tables = [t for t in tables if t in all_tables]
            missing = set(tables) - set(all_tables)
            if missing:
                result["warnings"] = [f"Tables not found: {', '.join(missing)}"]
        else:
            export_tables = all_tables

        # Export each table
        for table in export_tables:
            df = export_table_to_dataframe(conn, table)
            ws = wb.create_sheet(title=table[:31])  # Excel sheet name limit

            # Write data
            for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), 1):
                for c_idx, value in enumerate(row, 1):
                    ws.cell(row=r_idx, column=c_idx, value=value)

            if apply_format and len(df) > 0:
                apply_formatting(ws, df)

            result["sheets"].append({
                "name": table,
                "rows": len(df),
                "columns": len(df.columns),
            })
            result["total_rows"] += len(df)

        # Execute custom query if provided
        if custom_query:
            df = execute_query(conn, custom_query)
            ws = wb.create_sheet(title=query_name[:31])

            for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), 1):
                for c_idx, value in enumerate(row, 1):
                    ws.cell(row=r_idx, column=c_idx, value=value)

            if apply_format and len(df) > 0:
                apply_formatting(ws, df)

            result["sheets"].append({
                "name": query_name,
                "rows": len(df),
                "columns": len(df.columns),
                "query": custom_query,
            })
            result["total_rows"] += len(df)

        # Add schema sheet if requested
        if include_schema:
            create_schema_sheet(wb, conn, export_tables)
            result["sheets"].append({"name": "_Schema", "type": "schema"})

        wb.save(output_path)

    except Exception as e:
        result["success"] = False
        result["error"] = str(e)
    finally:
        conn.close()

    return result


def main():
    parser = argparse.ArgumentParser(description="Export SQLite database to Excel")
    parser.add_argument("database", help="Path to SQLite database file")
    parser.add_argument("--output", "-o", help="Output Excel file path")
    parser.add_argument("--tables", "-t", help="Comma-separated list of tables to export")
    parser.add_argument("--query", "-q", help="Custom SQL query to execute")
    parser.add_argument("--query-name", default="Query_Result", help="Sheet name for query result")
    parser.add_argument("--include-schema", "-s", action="store_true", help="Include schema sheet")
    parser.add_argument("--no-format", action="store_true", help="Skip formatting")

    args = parser.parse_args()

    db_path = Path(args.database)
    if not db_path.exists():
        print(f"Error: Database file not found: {db_path}", file=sys.stderr)
        sys.exit(1)

    output_path = Path(args.output) if args.output else db_path.with_suffix(".xlsx")
    tables = args.tables.split(",") if args.tables else None

    result = sqlite_to_xlsx(
        db_path=db_path,
        output_path=output_path,
        tables=tables,
        custom_query=args.query,
        query_name=args.query_name,
        include_schema=args.include_schema,
        apply_format=not args.no_format,
    )

    if result["success"]:
        print(f"✅ Exported to: {result['output_file']}")
        print(f"   Sheets: {len(result['sheets'])}")
        print(f"   Total rows: {result['total_rows']}")
        for sheet in result["sheets"]:
            if sheet.get("type") == "schema":
                print(f"   - {sheet['name']} (schema)")
            else:
                print(f"   - {sheet['name']}: {sheet['rows']} rows × {sheet['columns']} cols")
        if result.get("warnings"):
            for warning in result["warnings"]:
                print(f"   ⚠️  {warning}")
    else:
        print(f"❌ Error: {result['error']}", file=sys.stderr)
        sys.exit(1)


if __name__ == "__main__":
    main()
